# Reporting test for subhound.core.identify against the labeled
# portable_media_test_set.
#
# It runs identify() on every video file the manifest labels as a movie or a TV
# episode, classifies each result, and writes two human-readable artifacts:
#   tests/reports/identify_report.html
#   tests/reports/identify_report.xlsx
#
# Hard assertions (the rest is informational, because exact title matching needs
# online metadata for native-language/diacritic/expanded titles):
#   * The ONLY files allowed to be classified "unknown" are the 10 positional
#     "Fallen_Angel.mkv" episodes (no episode number anywhere in the name).
#   * No file may fail unexpectedly: for every non-unknown file the media type,
#     season and episode must match the manifest, and the year must match when
#     the manifest specifies one.
#
# Run just this test:  uv run pytest tests/test_identify_report.py -q

from __future__ import annotations

import csv
import datetime as dt
import html
import re
from dataclasses import dataclass
from pathlib import Path

import pytest

from subhound.core.identify import MOVIE, TV, UNKNOWN, identify
from subhound.core.scan import is_video_file
from tests.eval_identify import SHA256SUMS, verify_checksums

DATASET = Path(__file__).parent / "data" / "portable_media_test_set"
MANIFEST = DATASET / "manifest.csv"
REPORT_DIR = Path(__file__).parent / "reports"
KIND_TO_TYPE = {"movie": MOVIE, "episode": TV}

# Repo root (".../subhound"); used to rewrite absolute paths in the reports so
# they start at the repo name instead of leaking the machine-specific prefix.
REPO_ROOT = Path(__file__).resolve().parents[1]


# Function Summary:
#    Rewrite any absolute repo path in a string so it starts at the repo dir
#    name ("subhound/...") instead of the machine-specific absolute prefix.
#
#  Input (parameters):
#    text [str]:  any string that may embed an absolute path under the repo
#
#  Output:
#    text [str]:  the same string with the absolute repo prefix shortened
#
# Example:
#    repo_rel("/coding/subhound/tests/x.mkv")  ->  "subhound/tests/x.mkv"
def repo_rel(text: str) -> str:
  return (text or "").replace(str(REPO_ROOT), REPO_ROOT.name)

# The only files we expect to be undetermined: positional episodes with no
# number anywhere in the filename.
EXPECTED_UNKNOWN_BASENAME = "Fallen_Angel.mkv"

# Status -> (label, html color) used to colour both reports.
STATUS_STYLE = {
  "ok": ("OK", "C6EFCE"),
  "title_only": ("Title differs (canonicalization)", "FFEB9C"),
  "expected_unknown": ("Unknown (expected)", "BDD7EE"),
  "unexpected_unknown": ("Unknown (UNEXPECTED)", "FFC7CE"),
  "type_mismatch": ("Type mismatch", "FFC7CE"),
  "field_mismatch": ("Season/Episode/Year mismatch", "FFC7CE"),
}
FAILURE_STATUSES = {"unexpected_unknown", "type_mismatch", "field_mismatch"}


@dataclass
class Record:
  path: str
  kind: str
  exp_type: str
  exp_title: str
  exp_year: int | None
  exp_season: int | None
  exp_episode: int | None
  got_type: str
  got_title: str
  got_year: int | None
  got_season: int | None
  got_episode: int | None
  status: str
  note: str


# Function Summary:
#    Normalize a title for tolerant comparison (lowercase, drop non-alphanumerics,
#    collapse whitespace).
#
#  Input (parameters):
#    s [str]:  a title string
#
#  Output:
#    norm [str]:  the normalized comparison key
#
# Example:
#    _norm_title("The Office (US)")  ->  "the office us"
def _norm_title(s: str) -> str:
  s = re.sub(r"[^a-z0-9]+", " ", (s or "").lower())
  return re.sub(r"\s+", " ", s).strip()


# Function Summary:
#    Parse the first episode number from a manifest "episodes" cell ("", "5",
#    "1,2", ...).
#
#  Input (parameters):
#    cell [str]:  the manifest episodes value
#
#  Output:
#    ep [int | None]:  the first episode number, or None
#
# Example:
#    _first_episode("1,2")  ->  1
def _first_episode(cell: str) -> int | None:
  nums = re.findall(r"\d+", cell or "")
  return int(nums[0]) if nums else None


# Function Summary:
#    Coerce a possibly-blank manifest integer cell to int or None.
#
#  Input (parameters):
#    cell [str]:  a manifest cell expected to hold an int
#
#  Output:
#    value [int | None]:  the integer, or None when blank/non-numeric
#
# Example:
#    _as_int("2008")  ->  2008
def _as_int(cell: str) -> int | None:
  cell = (cell or "").strip()
  return int(cell) if cell.isdigit() else None


# Function Summary:
#    Classify a single identify() result against its ground-truth row, using the
#    test's leniency rules (title differences and manifest-absent years are not
#    failures).
#
#  Input (parameters):
#    rec_path [str]:        the manifest-relative path (to spot the expected unknowns)
#    info [MediaInfo]:      the identify() result
#    exp_type [str]:        expected media type
#    exp_title [str]:       expected title
#    exp_year [int|None]:   expected year (None if manifest is blank)
#    exp_season [int|None]: expected season
#    exp_episode [int|None]:expected episode
#
#  Output:
#    status [str]:  a key from STATUS_STYLE
#
# Example:
#    _classify("a/Fallen_Angel.mkv", info_unknown, "tv", "X", 1993, 1, 10)  ->  "expected_unknown"
def _classify(
  rec_path: str,
  info,
  exp_type: str,
  exp_title: str,
  exp_year: int | None,
  exp_season: int | None,
  exp_episode: int | None,
) -> str:
  is_expected_unknown = Path(rec_path).name == EXPECTED_UNKNOWN_BASENAME
  if info.media_type == UNKNOWN:
    return "expected_unknown" if is_expected_unknown else "unexpected_unknown"
  if info.media_type != exp_type:
    return "type_mismatch"
  if exp_type == TV and (info.season != exp_season or info.episode != exp_episode):
    return "field_mismatch"
  if exp_year is not None and info.year != exp_year:
    return "field_mismatch"
  if _norm_title(info.title_or_show) != _norm_title(exp_title):
    return "title_only"
  return "ok"


# Function Summary:
#    Run identify() over all manifest-labelled movie/episode video files and
#    build classified records.
#
#  Input (parameters):
#    (none)
#
#  Output:
#    records [list[Record]]:  one record per scored video file
#
# Example:
#    build_records()[0].status  ->  "ok"
def build_records() -> list[Record]:
  records: list[Record] = []
  for r in csv.DictReader(MANIFEST.open()):
    exp_type = KIND_TO_TYPE.get(r["kind"])
    if exp_type is None or not is_video_file(Path(r["path"])):
      continue
    exp_year = _as_int(r["year"])
    exp_season = _as_int(r["season"])
    exp_episode = _first_episode(r["episodes"])
    info = identify(DATASET / r["path"], None)
    status = _classify(r["path"], info, exp_type, r["expected_title"],
                       exp_year, exp_season, exp_episode)
    records.append(Record(
      path=r["path"], kind=r["kind"], exp_type=exp_type,
      exp_title=r["expected_title"], exp_year=exp_year, exp_season=exp_season,
      exp_episode=exp_episode, got_type=info.media_type,
      got_title=info.title_or_show, got_year=info.year, got_season=info.season,
      got_episode=info.episode, status=status, note=info.note,
    ))
  return records


# Function Summary:
#    Summarize records into counts per status plus per-field accuracy.
#
#  Input (parameters):
#    records [list[Record]]:  classified records
#
#  Output:
#    summary [dict]:  {"total", "status_counts", "field_acc"} aggregates
#
# Example:
#    summarize(records)["total"]  ->  333
def summarize(records: list[Record]) -> dict:
  status_counts: dict[str, int] = {k: 0 for k in STATUS_STYLE}
  fields = {"type": 0, "title": 0, "year": 0, "season": 0, "episode": 0}
  for rec in records:
    status_counts[rec.status] += 1
    fields["type"] += rec.got_type == rec.exp_type
    fields["title"] += _norm_title(rec.got_title) == _norm_title(rec.exp_title)
    fields["year"] += (rec.exp_year is None) or (rec.got_year == rec.exp_year)
    if rec.exp_type == TV:
      fields["season"] += rec.got_season == rec.exp_season
      fields["episode"] += rec.got_episode == rec.exp_episode
  tv_total = sum(1 for r in records if r.exp_type == TV) or 1
  total = len(records) or 1
  field_acc = {
    "type": fields["type"] / total,
    "title": fields["title"] / total,
    "year": fields["year"] / total,
    "season": fields["season"] / tv_total,
    "episode": fields["episode"] / tv_total,
  }
  return {"total": len(records), "status_counts": status_counts, "field_acc": field_acc}


# Function Summary:
#    Write the HTML report (data-integrity banner + summary cards + colour-coded
#    detail table).
#
#  Input (parameters):
#    records [list[Record]]:  classified records
#    summary [dict]:          the summarize() output
#    checks [dict]:           the checksum_status() output
#    out_path [Path]:         destination .html file
#
#  Output:
#    written [Path]:  the path written
#
# Example:
#    write_html(records, summary, checks, Path("r.html"))  ->  PosixPath("r.html")
def write_html(records: list[Record], summary: dict, checks: dict, out_path: Path) -> Path:
  def cell(v) -> str:
    return html.escape("" if v is None else str(v))

  if checks["ok"]:
    banner = (
      f'<div class="banner ok">&#10004; Data integrity verified &mdash; '
      f'{checks["verified"]}/{checks.get("total", checks["verified"])} files match '
      f'their SHA-256 checksums</div>'
    )
  else:
    detail = html.escape("; ".join(checks["problems"][:5]))
    banner = (
      f'<div class="banner bad">&#10006; DATA INTEGRITY FAILURE &mdash; '
      f'{len(checks["problems"])} checksum problem(s): {detail}</div>'
    )

  rows_html = []
  for rec in sorted(records, key=lambda r: (r.status not in FAILURE_STATUSES, r.path)):
    label, color = STATUS_STYLE[rec.status]
    rows_html.append(
      f'<tr style="background:#{color}">'
      f"<td>{cell(repo_rel(rec.path))}</td><td>{cell(rec.kind)}</td>"
      f"<td>{cell(rec.exp_type)}</td><td>{cell(rec.got_type)}</td>"
      f"<td>{cell(rec.exp_title)}</td><td>{cell(rec.got_title)}</td>"
      f"<td>{cell(rec.exp_year)}</td><td>{cell(rec.got_year)}</td>"
      f"<td>{cell(rec.exp_season)}</td><td>{cell(rec.got_season)}</td>"
      f"<td>{cell(rec.exp_episode)}</td><td>{cell(rec.got_episode)}</td>"
      f"<td>{html.escape(label)}</td><td>{cell(repo_rel(rec.note))}</td></tr>"
    )

  sc = summary["status_counts"]
  fa = summary["field_acc"]
  cards = "".join(
    f'<div class="card" style="background:#{STATUS_STYLE[k][1]}">'
    f'<div class="n">{sc[k]}</div><div class="l">{html.escape(STATUS_STYLE[k][0])}</div></div>'
    for k in STATUS_STYLE
  )
  field_rows = "".join(
    f"<tr><td>{k}</td><td>{v * 100:.1f}%</td></tr>" for k, v in fa.items()
  )
  generated = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
  doc = f"""<!doctype html><html><head><meta charset="utf-8">
<title>subhound identify report</title>
<style>
 body{{font-family:system-ui,Arial,sans-serif;margin:24px;color:#222}}
 h1{{margin:0 0 4px}} .sub{{color:#666;margin-bottom:16px}}
 .cards{{display:flex;flex-wrap:wrap;gap:10px;margin-bottom:18px}}
 .card{{border:1px solid #0002;border-radius:8px;padding:10px 14px;min-width:120px}}
 .card .n{{font-size:24px;font-weight:700}} .card .l{{font-size:12px}}
 table{{border-collapse:collapse;width:100%;font-size:12px}}
 th,td{{border:1px solid #0002;padding:4px 6px;text-align:left;vertical-align:top}}
 th{{background:#f3f3f3;position:sticky;top:0}}
 .acc{{width:auto;margin-bottom:18px}}
 .banner{{padding:10px 14px;border-radius:8px;font-weight:700;margin-bottom:16px}}
 .banner.ok{{background:#C6EFCE;color:#1b5e20;border:1px solid #1b5e2033}}
 .banner.bad{{background:#FFC7CE;color:#8b0000;border:1px solid #8b000033}}
</style></head><body>
<h1>subhound &mdash; identify() report</h1>
{banner}
<div class="sub">{summary['total']} video files scored against manifest ground truth &middot; generated {generated}</div>
<div class="cards">{cards}</div>
<table class="acc"><tr><th>Field</th><th>Accuracy</th></tr>{field_rows}</table>
<table>
<tr><th>Path</th><th>Kind</th><th>exp type</th><th>got type</th>
<th>exp title</th><th>got title</th><th>exp yr</th><th>got yr</th>
<th>exp S</th><th>got S</th><th>exp E</th><th>got E</th><th>Status</th><th>Note</th></tr>
{''.join(rows_html)}
</table></body></html>"""
  out_path.parent.mkdir(parents=True, exist_ok=True)
  out_path.write_text(doc, encoding="utf-8")
  return out_path


# Function Summary:
#    Write the Excel report: a Summary sheet and a colour-coded Details sheet.
#
#  Input (parameters):
#    records [list[Record]]:  classified records
#    summary [dict]:          the summarize() output
#    checks [dict]:           the checksum_status() output
#    out_path [Path]:         destination .xlsx file
#
#  Output:
#    written [Path]:  the path written
#
# Example:
#    write_xlsx(records, summary, checks, Path("r.xlsx"))  ->  PosixPath("r.xlsx")
def write_xlsx(records: list[Record], summary: dict, checks: dict, out_path: Path) -> Path:
  from openpyxl import Workbook
  from openpyxl.styles import Font, PatternFill
  from openpyxl.utils import get_column_letter

  wb = Workbook()
  ws = wb.active
  ws.title = "Summary"
  ws.append(["subhound identify() report"])
  ws["A1"].font = Font(bold=True, size=14)
  # Data-integrity banner row, colour-coded green (ok) / red (corrupted).
  if checks["ok"]:
    banner_text = (f"✔ Data integrity verified - {checks['verified']}/"
                   f"{checks.get('total', checks['verified'])} files match SHA-256 checksums")
    banner_color = STATUS_STYLE["ok"][1]
  else:
    banner_text = (f"✖ DATA INTEGRITY FAILURE - {len(checks['problems'])} "
                   f"checksum problem(s): " + "; ".join(checks["problems"][:5]))
    banner_color = STATUS_STYLE["type_mismatch"][1]
  ws.append([banner_text])
  ws.cell(ws.max_row, 1).font = Font(bold=True)
  ws.cell(ws.max_row, 1).fill = PatternFill("solid", fgColor=banner_color)
  ws.append(["Generated", dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
  ws.append(["Total scored", summary["total"]])
  ws.append([])
  ws.append(["Status", "Count"])
  for k in STATUS_STYLE:
    label, color = STATUS_STYLE[k]
    ws.append([label, summary["status_counts"][k]])
    ws.cell(ws.max_row, 1).fill = PatternFill("solid", fgColor=color)
  ws.append([])
  ws.append(["Field", "Accuracy"])
  for k, v in summary["field_acc"].items():
    ws.append([k, f"{v * 100:.1f}%"])

  ds = wb.create_sheet("Details")
  headers = ["Path", "Kind", "exp type", "got type", "exp title", "got title",
             "exp year", "got year", "exp S", "got S", "exp E", "got E",
             "Status", "Note"]
  ds.append(headers)
  for c in range(1, len(headers) + 1):
    ds.cell(1, c).font = Font(bold=True)
  ds.freeze_panes = "A2"
  for rec in sorted(records, key=lambda r: (r.status not in FAILURE_STATUSES, r.path)):
    label, color = STATUS_STYLE[rec.status]
    ds.append([
      repo_rel(rec.path), rec.kind, rec.exp_type, rec.got_type, rec.exp_title,
      rec.got_title, rec.exp_year, rec.got_year, rec.exp_season,
      rec.got_season, rec.exp_episode, rec.got_episode, label, repo_rel(rec.note),
    ])
    fill = PatternFill("solid", fgColor=color)
    for c in range(1, len(headers) + 1):
      ds.cell(ds.max_row, c).fill = fill
  widths = [60, 10, 9, 9, 26, 26, 8, 8, 6, 6, 6, 6, 30, 50]
  for i, w in enumerate(widths, start=1):
    ds.column_dimensions[get_column_letter(i)].width = w

  out_path.parent.mkdir(parents=True, exist_ok=True)
  wb.save(out_path)
  return out_path


# Function Summary:
#    Verify the dataset checksums and summarize the result for display in the
#    reports (so each report self-attests the integrity of the data it scored).
#
#  Input (parameters):
#    (none)
#
#  Output:
#    status [dict]:  {"ok" [bool], "verified" [int], "problems" [list[str]]}
#
# Example:
#    checksum_status()  ->  {"ok": True, "verified": 399, "problems": []}
def checksum_status() -> dict:
  if not SHA256SUMS.exists():
    return {"ok": False, "verified": 0, "problems": ["SHA256SUMS.txt not found"]}
  problems = verify_checksums(DATASET, SHA256SUMS)
  total = sum(1 for ln in SHA256SUMS.read_text(encoding="utf-8").splitlines() if ln.strip())
  return {"ok": not problems, "verified": total - len(problems), "problems": problems,
          "total": total}


@pytest.fixture(scope="module")
def report():
  if not MANIFEST.exists():
    pytest.skip(f"dataset not present at {MANIFEST}")
  checks = checksum_status()
  records = build_records()
  summary = summarize(records)
  write_html(records, summary, checks, REPORT_DIR / "identify_report.html")
  write_xlsx(records, summary, checks, REPORT_DIR / "identify_report.xlsx")
  return records, summary


def test_reports_written(report):
  # Both artifacts must exist and be non-trivial.
  assert (REPORT_DIR / "identify_report.html").stat().st_size > 0
  assert (REPORT_DIR / "identify_report.xlsx").stat().st_size > 0


def test_only_fallen_angel_is_unknown(report):
  # The ONLY files allowed to be undetermined are the positional Fallen_Angel
  # episodes (one per season-folder naming variant).
  records, _ = report
  unknown = [r.path for r in records if r.got_type == UNKNOWN]
  assert unknown, "expected at least the Fallen_Angel cases to be unknown"
  bad = [p for p in unknown if Path(p).name != EXPECTED_UNKNOWN_BASENAME]
  assert not bad, f"unexpected files classified as unknown: {bad}"


def test_no_unexpected_failures(report):
  # No structural mismatches (type/season/episode, and year when the manifest
  # provides it). Title differences are informational and allowed.
  records, _ = report
  failures = [
    f"{r.path} [{r.status}] got(type={r.got_type},S={r.got_season},"
    f"E={r.got_episode},yr={r.got_year})"
    for r in records if r.status in {"type_mismatch", "field_mismatch", "unexpected_unknown"}
  ]
  assert not failures, "unexpected identification failures:\n" + "\n".join(failures)
